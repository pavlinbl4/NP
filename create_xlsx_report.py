"""
проверя скачанные парсером снимки нового проспекта за месяц и генерю
файл отчета с превью
"""
from pathlib import Path
from openpyxl import Workbook

from datetime import datetime
from openpyxl.utils import get_column_letter
from folder_in_MY_documents import make_documets_folder
from openpyxl.styles import (
    Border, Side,
    Alignment, Font
)
from image_resize import image_resize
from set_column_dimensions import set_column_dimensions

filename_data = datetime.now().day
month_name = datetime.now().month
current_year = datetime.now().year

way_to_files = Path(
    f"{make_documets_folder('NewProspect')}/{current_year}_{month_name}")  # путь к папке с изображениями

workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Sheet with image"  # задаю название вкладки

# in list widths of all columns
worksheet = set_column_dimensions(worksheet, [20, 95, 40, 20])

thin_border = Border(left=Side(border_style="thin"),
                     right=Side(border_style="thin"),
                     top=Side(border_style="thin"),
                     bottom=Side(border_style="thin"))

for row, image_path in enumerate(way_to_files.glob("*.JPG"), 1):
    worksheet.row_dimensions[row].height = 150  # задаю высоту столбца
    print(image_path)

    # resize image
    img = image_resize(image_path)

    worksheet[f'{get_column_letter(3)}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet[f'{get_column_letter(3)}{row}'].border = thin_border
    worksheet.add_image(img, f"{get_column_letter(3)}{row}")

    worksheet[f'{get_column_letter(1)}{row}'].font = Font(size=14, bold=True)
    worksheet[f'{get_column_letter(1)}{row}'].alignment = Alignment(vertical='center')
    worksheet[f'{get_column_letter(1)}{row}'].border = thin_border
    worksheet[f'{get_column_letter(1)}{row}'].value = image_path.name.split("__")[0]

    worksheet[f'{get_column_letter(2)}{row}'].font = Font(size=12, bold=True)
    worksheet[f'{get_column_letter(2)}{row}'].alignment = Alignment(wrap_text=True, vertical='center')
    worksheet[f'{get_column_letter(2)}{row}'].border = thin_border
    worksheet[f'{get_column_letter(2)}{row}'].value = image_path.name.split("__")[1][:-4]

    worksheet[f'{get_column_letter(4)}{row}'].font = Font(size=14, bold=True)
    worksheet[f'{get_column_letter(4)}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet[f'{get_column_letter(4)}{row}'].border = thin_border
    worksheet[f'{get_column_letter(4)}{row}'].value = 500

worksheet[f'{get_column_letter(4)}{row + 3}'].alignment = Alignment(horizontal='center', vertical='center')
worksheet[f'{get_column_letter(4)}{row + 3}'].font = Font(size=17, bold=True, color="FF0000")
worksheet[f'{get_column_letter(4)}{row + 3}'].border = thin_border
worksheet[f'{get_column_letter(4)}{row + 3}'].value = f"=SUM(D$1:D${row})"

worksheet[f'{get_column_letter(2)}{row + 3}'].font = Font(size=17, bold=True, color="FF0000")
worksheet[f'{get_column_letter(2)}{row + 3}'].alignment = Alignment(horizontal='center', vertical='center')
worksheet[f'{get_column_letter(2)}{row + 3}'].border = thin_border
worksheet[f'{get_column_letter(2)}{row + 3}'].value = "ИТОГО"

workbook.save(f"{way_to_files}/report_{current_year}_{month_name}.xlsx")
